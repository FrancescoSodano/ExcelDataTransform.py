import os
import sys
import traceback
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import timedelta

def build_records_from_timesheet(timesheet_file, mapping_file):
    """
    Reads the timesheet file (possibly with multiple sheets) and mapping file,
    cleans and aggregates data by [DATA, SURNAME] (joining COMMESSA values and summing ORE).
    Surnames are converted to lower-case for case-insensitive matching.
    """
    # Read mapping file and create a dictionary for commessa mapping.
    df_map = pd.read_excel(mapping_file)
    commessa_map = dict(zip(df_map.iloc[:, 0], df_map.iloc[:, 1]))

    # Read all sheets from the timesheet file.
    sheets_dict = pd.read_excel(timesheet_file, sheet_name=None)
    df_raw_list = []
    for sheet_name, df_sheet in sheets_dict.items():
        if "Codice Commessa" not in df_sheet.columns:
            df_sheet["Codice Commessa"] = sheet_name
        df_raw_list.append(df_sheet)
    
    # FutureWarning may appear if some sheets are empty.
    df_raw = pd.concat(df_raw_list, ignore_index=True)

    # Define weekday offset map (Italian)
    day_offset = {
        "Lunedì": 0,
        "Martedì": 1,
        "Mercoledì": 2,
        "Giovedì": 3,
        "Venerdì": 4,
        "Sabato": 5,
        "Domenica": 6
    }

    records = []
    for idx, row in df_raw.iterrows():
        week_range_str = str(row.get("WeekRange", "")).strip()
        if " al " not in week_range_str:
            continue
        try:
            start_str, _ = week_range_str.split(" al ")
            start_date = pd.to_datetime(start_str, dayfirst=True)
        except Exception as e:
            print(f"Skipping row {idx} due to date parse error: {e}")
            continue

        # Map Codice Commessa using the mapping file.
        codice_commessa = row.get("Codice Commessa", "")
        commessa_final = commessa_map.get(codice_commessa, codice_commessa)

        # Extract surname from "Autore" (last word) and convert to lower-case.
        autore = str(row.get("Autore", "")).strip()
        surname = autore.split()[-1].lower() if autore else "unknown"

        # Process each weekday column.
        for day_col, offset in day_offset.items():
            hours = row.get(day_col, 0)
            # Clean hours value if it's a string (remove non-breaking spaces, extra whitespace).
            if isinstance(hours, str):
                cleaned_hours = hours.replace('\xa0', '').strip()
            else:
                cleaned_hours = hours
            try:
                hours_val = float(cleaned_hours)
            except Exception as e:
                hours_val = 0.0

            if pd.notna(hours) and hours_val != 0.0:
                actual_date = start_date + timedelta(days=offset)
                records.append({
                    "DATA": actual_date.date(),  # keep as date (original behavior)
                    "COMMESSA": commessa_final,
                    "ORE": hours_val,
                    "SURNAME": surname
                })

    if not records:
        print("No valid timesheet records found.")
        return pd.DataFrame()  # Return empty DataFrame if no records found.

    df_timesheet = pd.DataFrame(records)
    # Aggregate by DATA and SURNAME: join unique COMMESSA values and sum ORE.
    df_agg = df_timesheet.groupby(["DATA", "SURNAME"], as_index=False).agg({
        "COMMESSA": lambda x: "; ".join(sorted(set(x))),
        "ORE": "sum"
    })
    # Convert DATA to datetime for matching.
    df_agg["DATA"] = pd.to_datetime(df_agg["DATA"])
    return df_agg

def update_strategie_in_place(strategie_file, df_agg):
    """
    Updates the existing StrategieDigitali_2025 file in place.
    For each sheet (named for a surname), if a row's DATA value matches an aggregated date,
    update the COMMESSA and ORE cells.
    Surname matching is done in a case-insensitive way.
    """
    wb = openpyxl.load_workbook(strategie_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Compare using lower-case for case-insensitive matching.
        subset = df_agg[df_agg["SURNAME"].str.lower() == sheet_name.lower()].copy()
        if subset.empty:
            continue
        subset.set_index("DATA", inplace=True)

        # Iterate through rows of the sheet (assume headers in row 1 and columns A, B, C as DATA, COMMESSA, ORE).
        for row_cells in ws.iter_rows(min_row=2, max_col=3, values_only=False):
            date_cell, commessa_cell, ore_cell = row_cells

            cell_value = date_cell.value
            if isinstance(cell_value, str):
                try:
                    cell_value = pd.to_datetime(cell_value).date()
                except:
                    continue
            elif hasattr(cell_value, "date"):
                cell_value = cell_value.date()

            if cell_value in subset.index.date:
                match = subset.loc[subset.index.date == cell_value]
                if isinstance(match, pd.DataFrame) and len(match) > 1:
                    combined_commessa = "; ".join(sorted(set(match["COMMESSA"])))
                    combined_ore = match["ORE"].sum()
                else:
                    combined_commessa = match["COMMESSA"].values[0]
                    combined_ore = match["ORE"].values[0]
                commessa_cell.value = combined_commessa
                ore_cell.value = combined_ore

    wb.save(strategie_file)

def main():
    root = tk.Tk()
    root.withdraw()

    # Ask user for the timesheet file.
    timesheet_file = filedialog.askopenfilename(
        title="Select Timesheet Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if not timesheet_file:
        messagebox.showerror("Error", "No timesheet file selected. Exiting.")
        return

    # Ask user for the mapping file.
    mapping_file = filedialog.askopenfilename(
        title="Select Codice Commessa Mapping Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if not mapping_file:
        messagebox.showerror("Error", "No mapping file selected. Exiting.")
        return

    # Ask user for the StrategieDigitali_2025 file (to update in place).
    strategie_file = filedialog.askopenfilename(
        title="Select StrategieDigitali_2025 Excel File (to update in place)",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if not strategie_file:
        messagebox.showerror("Error", "No StrategieDigitali file selected. Exiting.")
        return

    # Build aggregated timesheet data.
    df_agg = build_records_from_timesheet(timesheet_file, mapping_file)
    if df_agg.empty:
        messagebox.showerror("Error", "No valid timesheet data found. Exiting.")
        return

    # Update the StrategieDigitali file in place.
    update_strategie_in_place(strategie_file, df_agg)
    messagebox.showinfo("Success", "Update complete. Your StrategieDigitali file has been updated.")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        root = tk.Tk()
        root.withdraw()
        error_details = traceback.format_exc()
        messagebox.showerror("Fatal Error", f"An unexpected error occurred:\n{error_details}")
